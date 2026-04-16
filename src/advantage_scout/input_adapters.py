from __future__ import annotations

import csv
from pathlib import Path
from typing import Any

from .schema import CANONICAL_HEADERS, HEADER_ALIASES, RawInput


def ingest_clipboard_text(text: str, source: str | None = "clipboard") -> RawInput:
    lines = [_clean_line(line) for line in text.splitlines() if line.strip()]
    if not lines:
        return RawInput(rows=[], source_type="raw clipboard text", source=source)

    rows = _parse_table(lines)
    if rows:
        return RawInput(
            rows=rows,
            source_type="raw clipboard text",
            source=source,
            headers=list(rows[0].keys()),
        )

    rows = _parse_flat_header_value_stream(lines)
    if rows:
        return RawInput(
            rows=rows,
            source_type="raw clipboard text",
            source=source,
            headers=list(rows[0].keys()),
        )

    key_value_row = _parse_key_value_lines(lines)
    if key_value_row:
        return RawInput(
            rows=[key_value_row],
            source_type="raw clipboard text",
            source=source,
            headers=list(key_value_row.keys()),
        )

    return RawInput(
        rows=[{"notes": "\n".join(lines)}],
        source_type="raw clipboard text",
        source=source,
        headers=["notes"],
        warnings=["could not detect a table or key/value rows; preserved input as notes"],
    )


def ingest_text_file(path: str | Path) -> RawInput:
    text_path = Path(path)
    raw_input = ingest_clipboard_text(text_path.read_text(encoding="utf-8"), source=str(text_path))
    raw_input.source_type = "raw text file"
    return raw_input


def ingest_xlsx(path: str | Path, sheet_name: str | None = None) -> RawInput:
    try:
        from openpyxl import load_workbook
    except ImportError as exc:
        raise RuntimeError("xlsx support requires openpyxl; install with `pip install -e .[xlsx]`") from exc

    xlsx_path = Path(path)
    workbook = load_workbook(xlsx_path, data_only=True, read_only=True)
    sheet = workbook[sheet_name] if sheet_name else workbook.active
    value_rows = [
        list(row)
        for row in sheet.iter_rows(values_only=True)
        if any(value is not None and str(value).strip() for value in row)
    ]
    if not value_rows:
        return RawInput(rows=[], source_type="xlsx file", source=str(xlsx_path))

    headers = _dedupe_headers([str(value).strip() if value is not None else "" for value in value_rows[0]])
    rows: list[dict[str, Any]] = []
    for values in value_rows[1:]:
        row = {
            header: value
            for header, value in zip(headers, values)
            if header and value is not None and str(value).strip()
        }
        if row:
            rows.append(row)

    return RawInput(rows=rows, source_type="xlsx file", source=str(xlsx_path), headers=headers)


def _parse_table(lines: list[str]) -> list[dict[str, str]]:
    for delimiter in ("|", "\t", ","):
        rows = _parse_delimited(lines, delimiter)
        if rows:
            return rows
    return _parse_whitespace_table(lines)


def _parse_flat_header_value_stream(lines: list[str]) -> list[dict[str, str]]:
    max_header_count = min(len(lines) // 2, len(CANONICAL_HEADERS))
    for header_count in range(max_header_count, 1, -1):
        header_tokens = lines[:header_count]
        if not _looks_like_header_block(header_tokens):
            continue

        value_tokens = lines[header_count:]
        if not value_tokens or len(value_tokens) % header_count != 0:
            continue

        rows = []
        for index in range(0, len(value_tokens), header_count):
            chunk = value_tokens[index : index + header_count]
            if len(chunk) < header_count:
                return []
            row = {
                header: value
                for header, value in zip(header_tokens, chunk)
                if header and value not in ("", None)
            }
            if row:
                rows.append(row)

        if rows:
            return rows

    return []


def _parse_delimited(lines: list[str], delimiter: str) -> list[dict[str, str]]:
    candidate_lines = [line for line in lines if delimiter in line]
    if len(candidate_lines) < 2:
        return []

    if delimiter == "|":
        split_lines = [
            [part.strip() for part in line.strip("|").split("|")]
            for line in candidate_lines
            if not _is_markdown_separator(line)
        ]
    else:
        split_lines = [[part.strip() for part in row] for row in csv.reader(candidate_lines, delimiter=delimiter)]

    return _rows_from_split_lines(split_lines)


def _parse_whitespace_table(lines: list[str]) -> list[dict[str, str]]:
    if len(lines) < 2 or not any("  " in line for line in lines):
        return []
    split_lines = [[part.strip() for part in line.split("  ") if part.strip()] for line in lines]
    return _rows_from_split_lines(split_lines)


def _rows_from_split_lines(split_lines: list[list[str]]) -> list[dict[str, str]]:
    if len(split_lines) < 2:
        return []

    headers = _dedupe_headers(split_lines[0])
    if len([header for header in headers if header]) < 2:
        return []

    rows: list[dict[str, str]] = []
    for values in split_lines[1:]:
        row = {
            header: value
            for header, value in zip(headers, values)
            if header and value not in ("", None)
        }
        if row:
            rows.append(row)
    return rows


def _parse_key_value_lines(lines: list[str]) -> dict[str, str]:
    row: dict[str, str] = {}
    for line in lines:
        separator = ":" if ":" in line else "=" if "=" in line else None
        if separator is None:
            continue
        key, value = line.split(separator, 1)
        if key.strip() and value.strip():
            row[key.strip()] = value.strip()
    return row


def _dedupe_headers(headers: list[str]) -> list[str]:
    counts: dict[str, int] = {}
    deduped_headers: list[str] = []
    for header in headers:
        clean_header = str(header).strip()
        if not clean_header:
            deduped_headers.append("")
            continue
        counts[clean_header] = counts.get(clean_header, 0) + 1
        deduped_headers.append(clean_header if counts[clean_header] == 1 else f"{clean_header}_{counts[clean_header]}")
    return deduped_headers


def _is_markdown_separator(line: str) -> bool:
    stripped = line.strip().strip("|").replace("|", "").replace(" ", "")
    return bool(stripped) and set(stripped) <= {"-", ":"}


def _looks_like_header_block(tokens: list[str]) -> bool:
    normalized = [_normalize_header_token(token) for token in tokens]
    if len(set(normalized)) != len(normalized):
        return False
    recognized = [token for token in tokens if _is_known_header(token)]
    return len(recognized) == len(tokens)


def _is_known_header(token: str) -> bool:
    normalized_token = _normalize_header_token(token)
    for alias in _header_alias_tokens():
        if normalized_token == alias:
            return True
        if normalized_token.startswith(alias):
            return True
        if alias in normalized_token:
            return True
    return False


def _header_alias_tokens() -> set[str]:
    aliases = {_normalize_header_token(header) for header in CANONICAL_HEADERS}
    for values in HEADER_ALIASES.values():
        aliases.update(_normalize_header_token(value) for value in values)
    return aliases


def _normalize_header_token(token: str) -> str:
    cleaned = token.strip().lower().replace("/", " per ")
    return " ".join("".join(character if character.isalnum() or character == "#" else " " for character in cleaned).split())


def _clean_line(line: str) -> str:
    return " ".join(line.replace("\u00a0", " ").split())
